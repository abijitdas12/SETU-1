"""
SETU Google Earth Engine (GEE) Disruption Hazard Sampler
=========================================================
Implements the exact Google Earth Engine Python API (`ee`) point-sampling pattern
to extract:
1. Rainfall (NASA GPM IMERG / ERA5)
2. Slope & Elevation (NASA SRTM 30m / ALOS DSM)
3. Vegetation NDVI (Copernicus Sentinel-2 MSI / MODIS 16-Day 250m)
4. Drainage Density (HydroSHEDS Global River Network)
5. Disruption Status (Ground truth incident or Hazard classification)

And fills any CSV (e.g., Bhukosh CSV or live lat/lon points) row by row,
exporting to standard CSV and Excel (.xlsx) formats.
"""

import os
import csv
import math
from datetime import datetime, timedelta
from typing import List, Dict, Any, Optional

# Safe Google Earth Engine (ee) import with explicit definition
try:
    # pyrefly: ignore [missing-import]
    import ee
    HAS_GEE = True
except ImportError:
    ee = None
    HAS_GEE = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def initialize_earth_engine(project_id: Optional[str] = None) -> bool:
    """
    Initializes Google Earth Engine API.
    Returns True if successful, False if credentials/init failed.
    """
    if not HAS_GEE or ee is None:
        raise ImportError(
            "Google Earth Engine Python API (`earthengine-api`) is not installed.\n"
            "Install it via: pip install earthengine-api"
        )
    try:
        if project_id:
            ee.Initialize(project=project_id)
        else:
            ee.Initialize()
        print("[+] Google Earth Engine initialized successfully.")
        return True
    except Exception as e:
        print(f"[-] GEE Authentication needed: {e}")
        print("[*] To authenticate, run in your terminal: earthengine authenticate")
        return False


class GEEHazardSampler:
    """
    Unified GEE Multi-Band Reducer for Coordinate Point Sampling.
    Uses `.reduceRegion()` pattern to extract climate, terrain, and vegetation features.
    """

    def __init__(self, project_id: Optional[str] = None):
        if not HAS_GEE or ee is None:
            raise ImportError(
                "Google Earth Engine library (`ee`) is not installed.\n"
                "Install it using: pip install earthengine-api"
            )

        # Attempt initialization if not already initialized
        try:
            # 1. Digital Elevation Model (SRTM 30m) & Computed Slope
            self.srtm = ee.Image("USGS/SRTMGL1_003")
            self.elevation = self.srtm.select("elevation")
            self.slope = ee.Terrain.slope(self.elevation).rename("slope")

            # 2. HydroSHEDS Global Drainage / River Network
            self.flow_accum = ee.Image("WWF/HydroSHEDS/15ACC").rename("drainage")
        except Exception:
            # If Earth Engine not initialized, attempt init
            initialize_earth_engine(project_id)
            self.srtm = ee.Image("USGS/SRTMGL1_003")
            self.elevation = self.srtm.select("elevation")
            self.slope = ee.Terrain.slope(self.elevation).rename("slope")
            self.flow_accum = ee.Image("WWF/HydroSHEDS/15ACC").rename("drainage")

    def sample_point_gee(
        self,
        lat: float,
        lon: float,
        date_str: str,
        buffer_meters: int = 100
    ) -> Dict[str, float]:
        """
        Samples GPM Rainfall, SRTM Slope/Elevation, Sentinel-2 NDVI, and Drainage
        at a specific (lat, lon, date) coordinate.
        """
        if not HAS_GEE or ee is None:
            raise RuntimeError("Earth Engine is not loaded.")

        point = ee.Geometry.Point([lon, lat])
        date_start = ee.Date(date_str)
        date_end = date_start.advance(1, "day")

        # --- A. Rainfall (NASA GPM IMERG 0.1° / 30-min to 24h accumulation) ---
        gpm_col = (
            ee.ImageCollection("NASA/GPM_L3/IMERG_V06")
            .filterDate(date_start, date_end)
            .select("precipitationCal")
        )
        # GPM gives precipitation rate (mm/hr); multiply by duration (0.5 hr per interval)
        rainfall_img = gpm_col.sum().multiply(0.5).rename("rainfall_24")

        # --- B. Vegetation NDVI (Sentinel-2 Level-2A with MODIS & baseline fallback) ---
        s2_start = date_start.advance(-7, "day")
        s2_end = date_start.advance(7, "day")

        s2_col = (
            ee.ImageCollection("COPERNICUS/S2_SR_HARMONIZED")
            .filterBounds(point)
            .filterDate(s2_start, s2_end)
            .filter(ee.Filter.lt("CLOUDY_PIXEL_PERCENTAGE", 30))
        )

        def add_ndvi(img):
            ndvi = img.normalizedDifference(["B8", "B4"]).rename("vegetation")
            return img.addBands(ndvi)

        s2_ndvi = s2_col.map(add_ndvi).select("vegetation").median()

        # Fallback MODIS NDVI if Sentinel-2 is cloud-covered
        modis_ndvi = (
            ee.ImageCollection("MODIS/061/MOD13Q1")
            .filterDate(date_start.advance(-16, "day"), date_start.advance(16, "day"))
            .select("NDVI")
            .median()
            .multiply(0.0001)
            .rename("vegetation")
        )

        default_veg = ee.Image.constant(0.55).rename("vegetation")
        modis_safe = modis_ndvi.unmask(default_veg)

        # Robust band merge preventing empty collection errors
        veg_img = ee.Image(
            ee.Algorithms.If(
                s2_col.size().gt(0),
                s2_ndvi.unmask(modis_safe),
                modis_safe
            )
        ).rename("vegetation")

        # --- C. Composite Multi-Band Image ---
        composite = (
            self.elevation
            .addBands(self.slope)
            .addBands(rainfall_img)
            .addBands(self.flow_accum)
            .addBands(veg_img)
        )

        # --- D. Single .reduceRegion() Call ---
        sampled_dict = composite.reduceRegion(
            reducer=ee.Reducer.mean(),
            geometry=point.buffer(buffer_meters),
            scale=30,
            maxPixels=1e8
        ).getInfo() or {}

        # Extract values with safe defaults
        rainfall_val = float(sampled_dict.get("rainfall_24") or 0.0)
        slope_val = float(sampled_dict.get("slope") or 0.0)
        elevation_val = float(sampled_dict.get("elevation") or 0.0)
        drainage_raw = float(sampled_dict.get("drainage") or 1.0)
        # Normalized drainage density index (log transformed flow accumulation)
        drainage_val = round(min(4.5, max(0.5, math.log10(max(1.0, drainage_raw)))), 2)
        veg_val = round(float(sampled_dict.get("vegetation") or 0.55), 2)

        return {
            "rainfall_24": round(rainfall_val, 2),
            "slope": round(slope_val, 2),
            "elevation": round(elevation_val, 1),
            "drainage": drainage_val,
            "vegetation": veg_val
        }


def export_to_excel(records: List[Dict[str, Any]], filepath: str):
    """Write records to professionally formatted Excel workbook (.xlsx)."""
    if not HAS_OPENPYXL:
        print("[-] openpyxl not installed. Skipping Excel export.")
        return

    os.makedirs(os.path.dirname(os.path.abspath(filepath)), exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "GEE Hazard Data"

    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    disrupted_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    disrupted_font = Font(name="Calibri", size=10, bold=True, color="991B1B")
    clear_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    clear_font = Font(name="Calibri", size=10, bold=True, color="166534")
    border_thin = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    headers = [
        "lat", "lon", "date", "rainfall_24", "slope", "drainage", "vegetation", "disruption",
        "district", "location_basis", "elevation_m"
    ]
    ws.append(headers)

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_thin

    for row_idx, r in enumerate(records, start=2):
        row_data = [
            r["lat"], r["lon"], r["date"], r["rainfall_24"], r["slope"],
            r["drainage"], r["vegetation"], r["disruption"],
            r.get("district", ""), r.get("location_basis", ""), r.get("elevation_m", 0.0)
        ]
        ws.append(row_data)

        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = border_thin
            cell.alignment = Alignment(horizontal="center" if col_idx <= 8 else "left", vertical="center")

            if col_idx == 8:
                if r["disruption"] == 1:
                    cell.fill = disrupted_fill
                    cell.font = disrupted_font
                else:
                    cell.fill = clear_fill
                    cell.font = clear_font

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    wb.save(filepath)
    print(f"[+] Successfully exported Excel workbook to: {filepath}")


def process_bhukosh_csv_with_gee(
    input_csv_path: str,
    output_csv_path: str,
    output_xlsx_path: Optional[str] = None
):
    """
    Takes an input Bhukosh CSV (with lat, lon, date), queries GEE for each row,
    and writes out standard columns: lat, lon, date, rainfall_24, slope, drainage, vegetation, disruption.
    """
    if not HAS_GEE or ee is None:
        print("[-] GEE Python package (`earthengine-api`) is not installed. Run: pip install earthengine-api")
        return

    sampler = GEEHazardSampler()
    rows_out = []

    with open(input_csv_path, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for i, row in enumerate(reader):
            lat = float(row["lat"])
            lon = float(row["lon"])
            date_str = row.get("date", datetime.now().strftime("%Y-%m-%d"))

            print(f"[*] Processing row {i+1}: ({lat}, {lon}) on {date_str}...")
            sampled = sampler.sample_point_gee(lat, lon, date_str)

            # Check disruption (ground truth or threshold)
            disrupted_raw = row.get("disrupted", row.get("disruption"))
            if disrupted_raw is not None:
                disruption_val = int(disrupted_raw)
            else:
                # Compound physical threshold
                disruption_val = 1 if (sampled["slope"] >= 15.0 and sampled["rainfall_24"] >= 60.0) else 0

            record = {
                "lat": lat,
                "lon": lon,
                "date": date_str,
                "rainfall_24": sampled["rainfall_24"],
                "slope": sampled["slope"],
                "drainage": sampled["drainage"],
                "vegetation": sampled["vegetation"],
                "disruption": disruption_val,
                "district": row.get("district", ""),
                "location_basis": row.get("location_basis", ""),
                "elevation_m": sampled["elevation"]
            }
            rows_out.append(record)

    # Save to CSV
    os.makedirs(os.path.dirname(os.path.abspath(output_csv_path)), exist_ok=True)
    fieldnames = ["lat", "lon", "date", "rainfall_24", "slope", "drainage", "vegetation", "disruption", "district", "location_basis", "elevation_m"]
    with open(output_csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for r in rows_out:
            writer.writerow(r)

    print(f"[+] Successfully generated GEE-sampled CSV: {output_csv_path}")

    # Save to Excel if requested
    if output_xlsx_path and HAS_OPENPYXL:
        export_to_excel(rows_out, output_xlsx_path)


if __name__ == "__main__":
    print("[*] GEE Hazard Sampler Module loaded.")
